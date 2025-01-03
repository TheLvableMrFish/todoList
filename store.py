from openpyxl import Workbook

def write_to_excel(tasks):

    wb = Workbook()
    ws = wb.active
    ws.title = "Tasks"
    
    titles = ['Task', 'Status', 'Date Added', 'Complete Date']

    for i, title in enumerate(titles):
        ws.cell(row=1, column= i + 1, value= title)

    for i, task in enumerate(tasks):
        for j, section in enumerate(task):
            ws.cell(row = i + 2, column = j + 1, value= section)

    wb.save("tasks.xlsx")


todoList = [
    ["task 1", "complete", "yesterday", "today"],
    ["task 2", "complete", "yesterday", "today"],
    ["task 3", "complete", "yesterday", "today"],
    ["task 4", "complete", "yesterday", "today"],
]