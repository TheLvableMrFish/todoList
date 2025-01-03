from openpyxl import load_workbook

def getTodoList():

    # Check if tasks exel file exists with the todo list
    try:
        row_values = []

        wb = load_workbook("tasks.xlsx")
        ws = wb.active

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            row_values.append([cell.value for cell in row])
    
    # If not return a blank array to start the todo list
    except FileNotFoundError:
        row_values = []

    return row_values